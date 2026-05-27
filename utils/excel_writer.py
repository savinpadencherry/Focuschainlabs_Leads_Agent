"""
Ranked lead workbook writer.

Output is a single 'Leads' sheet with a frozen header, score-based row
fills, and a deliberately limited column set so it's outreach-ready
when opened in Excel or Google Sheets.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


COLUMNS = [
    "Rank", "Company", "Website", "Location", "Industry",
    "Total Score", "Fit", "Trigger", "Reachability", "Recency",
    "Primary Signal", "Pain Point", "Score Reasoning",
    "Ad Activity", "Evidence",
    "Contact Name", "Contact Title", "Email", "LinkedIn URL",
    "Contact Posts", "Reddit Signals",
    "Opening Line", "Outreach Strategy",
    "Source", "Date Found", "Status",
]


def write_leads_to_excel(leads: list, output_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Brand-aligned header (ink #0F2A33, cream text)
    header_fill = PatternFill("solid", fgColor="0F2A33")
    header_font = Font(bold=True, color="F4F0E7", size=11, name="Inter")
    thin_border = Border(
        left=Side(style="thin", color="E5E0D3"),
        right=Side(style="thin", color="E5E0D3"),
        top=Side(style="thin", color="E5E0D3"),
        bottom=Side(style="thin", color="E5E0D3"),
    )

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26

    leads_sorted = sorted(leads, key=lambda x: x.get("total_score", 0), reverse=True)

    for rank, lead in enumerate(leads_sorted, 1):
        row = rank + 1
        score = lead.get("total_score", 0)

        if score >= 80:
            fill = PatternFill("solid", fgColor="E6F4EA")   # mint
        elif score >= 60:
            fill = PatternFill("solid", fgColor="F4F0E7")   # cream
        else:
            fill = PatternFill("solid", fgColor="FBFAF7")   # off-white

        contact_posts  = " || ".join(lead.get("contact_posts",  []) or [])[:600]
        reddit_signals = " || ".join(lead.get("reddit_signals", []) or [])[:600]

        # Ad activity — join detected signals
        ad_signals = lead.get("ad_signals", []) or []
        ad_activity = " | ".join(ad_signals[:3]) if ad_signals else ""

        # Evidence — structured items as readable lines
        evidence_items = lead.get("evidence", []) or []
        evidence_text = "\n".join(
            f"[{e.get('category','').upper()}] {e.get('observation','')[:120]}"
            + (f"\n  → {e['url']}" if e.get("url") else "")
            for e in evidence_items[:8]
        )

        values = [
            rank,
            lead.get("company_name", ""),
            lead.get("website", ""),
            (lead.get("locations", ["Bangalore"])[0]
                if isinstance(lead.get("locations"), list) else "Bangalore"),
            lead.get("vertical", ""),
            score,
            lead.get("fit_score", ""),
            lead.get("trigger_score", ""),
            lead.get("reachability_score", ""),
            lead.get("intent_recency_score", ""),
            lead.get("primary_signal", ""),
            lead.get("pain_point", ""),
            lead.get("score_reasoning", ""),
            ad_activity,
            evidence_text,
            lead.get("contact_name", ""),
            lead.get("contact_title", ""),
            lead.get("email", ""),
            lead.get("linkedin_url", ""),
            contact_posts,
            reddit_signals,
            lead.get("opening_line", ""),
            lead.get("outreach_note", ""),
            lead.get("source", ""),
            lead.get("date_found", ""),
            "New",
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.fill = fill
            cell.font = Font(name="Inter", size=10, color="0F2A33")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border

    # Auto-fit (capped) widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"  Saved: {output_path}")
    return output_path
